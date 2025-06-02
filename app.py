
from flask import Flask, render_template, request, redirect, url_for, send_file, session, Response, jsonify
from dotenv import load_dotenv
import psycopg2
import csv
import json
import io
import os
import re
from datetime import datetime, timedelta, timezone, timedelta   
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import tempfile
from datetime import datetime
import hashlib

# 🔄 Carregar variáveis do .env
load_dotenv()
print(os.getenv('DATABASE_URL'))

# 🔥 Instanciação da aplicação
app = Flask(__name__)

# 🔑 Definição da chave secreta (agora segura)
app.secret_key = os.getenv('SECRET_KEY')


# 🔗 Conexão com o banco de dados (segura)
def conectar():
    return psycopg2.connect(os.getenv('DATABASE_URL'))

VERSAO_ATUAL = "v1.4"

CONFIG_FILE = 'config.json'

def carregar_config():
    config_padrao = {
        "formulario_ativo": True,
        "audicao_ativa": True
    }

    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as file:
            try:
                config_arquivo = json.load(file)
                # Atualizar config_padrao com o que tiver no arquivo
                config_padrao.update(config_arquivo)
            except json.JSONDecodeError:
                # Arquivo corrompido, volta para padrão
                pass

    return config_padrao

def salvar_config(config):
    with open(CONFIG_FILE, 'w') as file:
        json.dump(config, file, indent=4)


@app.context_processor
def inject_version():
    return dict(versao=VERSAO_ATUAL)

def criptografar_senha(senha):
    return hashlib.sha256(senha.encode('utf-8')).hexdigest()


def admin_required(func):
    from functools import wraps
    @wraps(func)
    def decorated_view(*args, **kwargs):
        if 'usuario' not in session or session.get('nivel') != 'Admin':
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    return decorated_view

def calcular_idade(data_nascimento):
    if data_nascimento:
        nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d')
        hoje = datetime.today()
        idade = hoje.year - nascimento.year - ((hoje.month, hoje.day) < (nascimento.month, nascimento.day))
        return idade
    return None


# ===========================================
# 🔐 Menu Principal após Login
# ===========================================
@app.route('/menu')
def menu():
    if 'usuario' not in session:
        return redirect('/login')

    acesso_negado = request.args.get('acesso_negado') == '1'
    return render_template('menu.html', acesso_negado=acesso_negado)


# ===========================================
# 🔍 Verifica se o cadastro já existe (usado no formulário)
# ===========================================
@app.route('/verificar_cadastro', methods=['POST'])
def verificar_cadastro():
    data = request.get_json()
    nome = data.get('nome').strip().upper()
    nascimento = data.get('nascimento').strip()

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM inscritos WHERE UPPER(nome) = %s AND nascimento = %s", (nome, nascimento))
    resultado = cursor.fetchone()

    if resultado:
        keys = [desc[0] for desc in cursor.description]
        dados = dict(zip(keys, resultado))
        conn.close()
        return jsonify({"existe": True, "dados": dados})
    else:
        conn.close()
        return jsonify({"existe": False})


    

# ===========================================
# ⚠️ Página que informa que o cadastro já existe
# ===========================================
@app.route('/cadastro_existente')
def cadastro_existente():
    nome = request.args.get('nome', 'Usuário')  # Se não vier nome, mostra 'Usuário'
    return render_template('cadastro_existente.html', nome=nome)

# ===============================
# Login e Controle de Sessão
# ===============================


# ===========================================
# 🔑 Login de usuários e controle de sessão
# ===========================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    print("🚀 LOGIN acionado")
    erro = None
    if request.method == 'POST':
        usuario = request.form['usuario'].strip().lower()
        senha = request.form['senha'].strip()

        print(f"🧠 Usuario digitado: {usuario}")
        print(f"🧠 Senha digitada (criptografada): {criptografar_senha(senha)}")

        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("SELECT senha, nivel FROM usuarios WHERE usuario = %s", (usuario,))
        resultado = cursor.fetchone()
        conn.close()

        if resultado:
            senha_correta, nivel = resultado

            print(f"🔑 Senha no banco: {senha_correta}")

            if criptografar_senha(senha) == senha_correta.strip():
                session['usuario'] = usuario
                session['nivel'] = nivel
                return redirect('/menu')
            else:
                erro = "Senha incorreta!"
        else:
            erro = "Usuário não encontrado!"

    return render_template('login.html', erro=erro)




# ===========================================
# 🔒 Logout - Finaliza sessão
# ===========================================
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

# ===============================
# Cadastro e Visualizações
# ===============================


# ===========================================
# 📝 Formulário de cadastro e edição
# ===========================================
@app.route('/', methods=['GET', 'POST'])
def form():
    config = carregar_config()
    if not config.get('formulario_ativo', True):
        return render_template('form.html', mostrar_modal=True)
    sucesso = False
    erro = False
    if request.method == 'POST':
        try:
            dados = request.form.to_dict(flat=True)
            dados['nome'] = dados.get('nome', '').upper()
            dados['email'] = dados.get('email', '').lower()

            campos_title = ['endereco', 'bairro', 'cidade', 'estado', 'profissao']
            for campo in campos_title:
                if campo in dados and dados[campo]:
                    dados[campo] = dados[campo].title()

            conn = conectar()
            cursor = conn.cursor()

            if dados.get('id'):  # Se existe ID, faz UPDATE (alteração)
                cursor.execute('''
                    UPDATE inscritos SET
                        nome=%s, nascimento=%s, telefone=%s, email=%s, endereco=%s, numero=%s, complemento=%s,
                        bairro=%s, cidade=%s, estado=%s, profissao=%s, estado_civil=%s, batizado=%s, perfil=%s,
                        voz=%s, aulas=%s, segunda_voz=%s, instrumentos=%s, outro_instrumento=%s, cifras=%s, ouvido=%s,
                        nivel_tecnico=%s, ensaios=%s, cultos=%s, motivacao=%s, experiencia=%s
                    WHERE id=%s
                ''', (
                    dados.get('nome'), dados.get('nascimento'), dados.get('telefone'), dados.get('email'),
                    dados.get('endereco'), dados.get('numero'), dados.get('complemento'), dados.get('bairro'),
                    dados.get('cidade'), dados.get('estado'), dados.get('profissao'), dados.get('estado_civil'),
                    dados.get('batizado'), dados.get('perfil'), dados.get('voz'), dados.get('aulas'),
                    dados.get('segunda_voz'), dados.get('instrumentos'), dados.get('outro_instrumento'),
                    dados.get('cifras'), dados.get('ouvido'), dados.get('nivel_tecnico'),
                    dados.get('ensaios'), dados.get('cultos'), dados.get('motivacao'), dados.get('experiencia'),
                    dados.get('id')  # ID vai na cláusula WHERE
                ))
            else:  # Se não existe ID, faz INSERT (novo cadastro)
                cursor.execute('''
                    INSERT INTO inscritos (
                        nome, nascimento, telefone, email, endereco, numero, complemento,
                        bairro, cidade, estado, profissao, estado_civil, batizado, perfil,
                        voz, aulas, segunda_voz, instrumentos, outro_instrumento, cifras, ouvido,
                        nivel_tecnico, ensaios, cultos, motivacao, experiencia
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    dados.get('nome'), dados.get('nascimento'), dados.get('telefone'), dados.get('email'),
                    dados.get('endereco'), dados.get('numero'), dados.get('complemento'), dados.get('bairro'),
                    dados.get('cidade'), dados.get('estado'), dados.get('profissao'), dados.get('estado_civil'),
                    dados.get('batizado'), dados.get('perfil'), dados.get('voz'), dados.get('aulas'),
                    dados.get('segunda_voz'), dados.get('instrumentos'), dados.get('outro_instrumento'),
                    dados.get('cifras'), dados.get('ouvido'), dados.get('nivel_tecnico'),
                    dados.get('ensaios'), dados.get('cultos'), dados.get('motivacao'), dados.get('experiencia')
                ))

            conn.commit()
            conn.close()
            return redirect('/sucesso')
        except Exception as e:
            print(e)
            erro = True


    return render_template('form.html', sucesso=sucesso, erro=erro)


# ===========================================
# 📝 Pagina de formulario enviado
# ===========================================
@app.route('/sucesso')
def sucesso():
    return render_template('sucesso.html')


# ===========================================
# 📝 Pagina inscritos
# ===========================================
@app.route('/inscritos')
def inscritos():
    if 'usuario' not in session:
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM inscritos ORDER BY nome ASC')
    inscritos = cursor.fetchall()
    conn.close()

    inscritos_formatados = []

    for inscrito in inscritos:
        inscrito = list(inscrito)

        # Calcular idade de forma segura
        try:
            if inscrito[2] is not None:
                idade = calcular_idade(str(inscrito[2]))
            else:
                idade = None
        except Exception:
            idade = None

        inscrito.append(idade)

        # Formatar data de nascimento de forma segura
        try:
            if inscrito[2] is not None:
                data = datetime.strptime(str(inscrito[2]), "%Y-%m-%d")
                inscrito[2] = data.strftime("%d/%m/%Y")
            else:
                inscrito[2] = "Data não informada"
        except Exception:
            inscrito[2] = "Data inválida"

        inscritos_formatados.append(inscrito)

    return render_template('inscritos.html', inscritos=inscritos_formatados)





# ===========================================
# 📝 Pagina gerenciar
# ===========================================
@app.route('/gerenciar')
def gerenciar():
    if 'usuario' not in session:
        return redirect('/login')

    if session.get('nivel') != 'admin':
        return redirect(url_for('menu', acesso_negado='1'))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM inscritos ORDER BY nome ASC')
    inscritos = cursor.fetchall()
    conn.close()
    return render_template('gerenciar.html', inscritos=inscritos)



# ===========================================
# 📝 Pagina para acesso negado usuario
# ===========================================
@app.route('/acesso_negado')
def acesso_negado():
    if 'usuario' not in session:
        return redirect('/login')
    return render_template('acesso_negado.html')


# ===========================================
# 📝 Rota ajax ativa/desativar cadastro
# ===========================================
@app.route('/atualizar_status', methods=['POST'])
def atualizar_status():
    data = request.get_json()
    id = data.get('id')
    status = data.get('status')  # Boolean: true ou false

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('UPDATE inscritos SET ativo = %s WHERE id = %s', (status, id))
    conn.commit()
    conn.close()

    return jsonify({'success': True})




# ===========================================
# 📝 Rota inativar cadastro
# ===========================================
@app.route('/inativar/<int:id>')
def inativar(id):
    if 'usuario' not in session or session['nivel'] != 'admin':
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('UPDATE inscritos SET ativo = FALSE WHERE id = %s', (id,))
    conn.commit()
    conn.close()
    return redirect('/gerenciar')


# ===========================================
# 📝 Rota Ativar Cadastro
# ===========================================
@app.route('/ativar/<int:id>')
def ativar(id):
    if 'usuario' not in session or session['nivel'] != 'admin':
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('UPDATE inscritos SET ativo = TRUE WHERE id = %s', (id,))
    conn.commit()
    conn.close()
    return redirect('/gerenciar')


# ===========================================
# 📝 Excluir cadastros
# ===========================================
@app.route('/excluir/<int:id>')
def excluir(id):
    if 'usuario' not in session or session['nivel'] != 'admin':
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM inscritos WHERE id = %s', (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('gerenciar'))


# ===========================================
# 📝 Formulário edição de cadastros
# ===========================================
@app.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    if 'usuario' not in session or session['nivel'] != 'admin':
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()

    if request.method == 'POST':
        dados = request.form.to_dict(flat=True)
        dados['nome'] = dados.get('nome', '').upper()
        dados['email'] = dados.get('email', '').lower()

        campos_title = ['endereco', 'bairro', 'cidade', 'estado', 'profissao']
        for campo in campos_title:
            if campo in dados and dados[campo]:
                dados[campo] = dados[campo].title()

        cursor.execute('''
            UPDATE inscritos SET
                nome=%s, nascimento=%s, telefone=%s, email=%s, endereco=%s, numero=%s, complemento=%s,
                bairro=%s, cidade=%s, estado=%s, profissao=%s, estado_civil=%s, batizado=%s, perfil=%s,
                voz=%s, aulas=%s, segunda_voz=%s, instrumentos=%s, outro_instrumento=%s, cifras=%s, ouvido=%s,
                nivel_tecnico=%s, ensaios=%s, cultos=%s, motivacao=%s, experiencia=%s
            WHERE id=%s
        ''', (
            dados.get('nome'), dados.get('nascimento'), dados.get('telefone'), dados.get('email'),
            dados.get('endereco'), dados.get('numero'), dados.get('complemento'), dados.get('bairro'),
            dados.get('cidade'), dados.get('estado'), dados.get('profissao'), dados.get('estado_civil'),
            dados.get('batizado'), dados.get('perfil'), dados.get('voz'), dados.get('aulas'),
            dados.get('segunda_voz'), dados.get('instrumentos'), dados.get('outro_instrumento'),
            dados.get('cifras'), dados.get('ouvido'), dados.get('nivel_tecnico'),
            dados.get('ensaios'), dados.get('cultos'), dados.get('motivacao'), dados.get('experiencia'),
            id
        ))
        conn.commit()
        conn.close()
        return redirect(url_for('gerenciar'))

    cursor.execute('SELECT * FROM inscritos WHERE id = %s', (id,))
    inscrito = cursor.fetchone()
    conn.close()
    return render_template('editar.html', inscrito=inscrito)



# ===========================================
# 📝 Formulário detalhes do cadastro
# ===========================================
@app.route('/detalhes/<int:id>')
def detalhes(id):
    if 'usuario' not in session:
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM inscritos WHERE id = %s', (id,))
    inscrito = list(cursor.fetchone())
    conn.close()

    try:
        if inscrito[2]:
            data = datetime.strptime(str(inscrito[2]), "%Y-%m-%d")
            inscrito[2] = data.strftime("%d/%m/%Y")
    except Exception:
        inscrito[2] = "Data inválida"

    return render_template('detalhes.html', inscrito=inscrito)



# ===========================================
# 📝 Exporta CVS - NAO ESTA SENDO UTILIZADA
# ===========================================
@app.route('/exportar')
def exportar():
    if 'usuario' not in session:
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM inscritos ORDER BY nome ASC')
    inscritos = cursor.fetchall()
    conn.close()

    with open('inscritos.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Nome', 'Nascimento', 'Telefone', 'Email', 'Endereço', 'Número', 'Complemento',
                         'Bairro', 'Cidade', 'Estado', 'Profissão', 'Estado Civil', 'Batizado', 'Perfil',
                         'Classificação Vocal', 'Aulas', 'Segunda Voz', 'Instrumentos', 'Outro Instrumento',
                         'Sabe Cifras', 'Ouvido', 'Nível Técnico', 'Ensaios', 'Cultos', 'Motivação', 'Experiência', 'Status'])
        for i in inscritos:
            writer.writerow(i[1:-1])

    return send_file('inscritos.csv', as_attachment=True)


# ===========================================
# 📝 Exporta xlsx
# ===========================================
@app.route('/exportar_csv')
def exportar_xlsx():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from datetime import datetime
    import os

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM inscritos ORDER BY nome ASC')
    inscritos = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inscritos"

    # Título e subtítulo
    ws.merge_cells('A1:Y1')
    ws['A1'] = 'PENIEL GUILHERMINA'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='left')

    ws.merge_cells('A2:Y2')
    ws['A2'] = 'Cadastro Ministério de Louvor'
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='left')

    # Cabeçalhos
    cabecalhos = [
        'Nome', 'Nascimento', 'Telefone', 'Email', 'Endereço', 'Número', 'Complemento',
        'Bairro', 'Cidade', 'Estado', 'Profissão', 'Estado Civil', 'Batizado', 'Perfil',
        'Classificação Vocal', 'Aulas de Canto', 'Segunda Voz', 'Instrumentos', 'Cifras',
        'Ouvido', 'Nível Técnico', 'Ensaios', 'Cultos', 'Motivação', 'Experiência', 'Status'
    ]

    verde = '6FCF97'
    fill_cabecalho = PatternFill(start_color=verde, end_color=verde, fill_type='solid')
    borda_cabecalho = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    for col_num, header in enumerate(cabecalhos, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill_cabecalho
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borda_cabecalho

    # Inserção dos dados
    for row_num, inscrito in enumerate(inscritos, start=4):
        dados = [
            inscrito[1], inscrito[2], inscrito[3], inscrito[4], inscrito[5], inscrito[6],
            inscrito[7], inscrito[8], inscrito[9], inscrito[10], inscrito[11], inscrito[12],
            inscrito[13], inscrito[14], inscrito[15], inscrito[16], inscrito[17], inscrito[18],
            inscrito[19], inscrito[20], inscrito[21], inscrito[22], inscrito[23], inscrito[24],
            inscrito[25], 'Ativo' if inscrito[27] else 'Inativo'
        ]
        for col_num, dado in enumerate(dados, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = dado
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Largura de colunas personalizada
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'A':  # Nome
            ws.column_dimensions[col_letter].width = 35
        elif col_letter in ['X', 'Y']:  # Motivação e Experiência
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = min(30, max(len(str(cell.value)) for cell in col if cell.value) + 2)

    # Altura padronizada
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 40

    # Filtro apenas sobre a coluna "Nome" (coluna A)
    ultima_linha = ws.max_row
    ws.auto_filter.ref = f"A3:A{ultima_linha}"

    # Nome do arquivo com data
    data_str = datetime.now().strftime('%d%m%Y')
    nome_arquivo = f'Cadastro_Louvor_{data_str}.xlsx'

    # Salvar e enviar
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.seek(0)

    return send_file(temp_file.name,
                     as_attachment=True,
                     download_name=nome_arquivo,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



# ===========================================
# 📝 Pagina Senhas
# ===========================================
@app.route('/senhas')
def senhas():
    if 'usuario' not in session:
        return redirect('/login')

    if session.get('nivel') != 'admin':
        return redirect(url_for('menu', acesso_negado='1'))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usuarios")
    usuarios = cursor.fetchall()
    conn.close()
    return render_template('senhas.html', usuarios=usuarios)




# ===========================================
# 📝 Cadastros de usuarios
# ===========================================
@app.route('/novo_usuario', methods=['GET', 'POST'])
def novo_usuario():
    if 'usuario' not in session or session.get('nivel') != 'admin':
        return redirect('/login')

    if request.method == 'POST':
        usuario = request.form.get('usuario', '').strip().lower()
        senha = request.form.get('senha', '').strip()
        nivel = request.form.get('nivel', '').strip()

        if not usuario or not senha:
            mensagem = '❌ Usuário e senha são obrigatórios.'
            return render_template('novo_usuario.html', mensagem=mensagem)

        senha_cripto = criptografar_senha(senha)

        conn = conectar()
        cursor = conn.cursor()

        # Verifica se já existe
        cursor.execute("SELECT * FROM usuarios WHERE usuario = %s", (usuario,))
        existente = cursor.fetchone()

        if existente:
            conn.close()
            mensagem = f'❌ O usuário "{usuario}" já existe. Escolha outro nome.'
            return render_template('novo_usuario.html', mensagem=mensagem)
        else:
            cursor.execute("INSERT INTO usuarios (usuario, senha, nivel) VALUES (%s, %s, %s)",
                           (usuario, senha_cripto, nivel))
            conn.commit()
            conn.close()
            return redirect(url_for('senhas'))  # ✔️ Redireciona para a página senhas

    return render_template('novo_usuario.html')


# ===========================================
# 📝 Rota para excluir usuarios
# ===========================================
@app.route('/excluir_usuario/<int:id>')
def excluir_usuario(id):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM usuarios WHERE id = %s", (id,))
    cursor.execute("DELETE FROM usuarios WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('senhas'))


# ===========================================
# 📝 Formulário de cadastro e edição
# ===========================================
@app.route('/adicionar_campo', methods=['POST'])
def adicionar_campo():
    if request.method == 'POST':
        nome_campo = request.form.get('nova_pergunta')
        tipo_resposta = request.form.get('tipo_resposta')

        if nome_campo:
            adicionar_coluna_se_necessario(nome_campo.replace(' ', '_').lower())

        return redirect('/layout')


# ===========================================
# 📝 Rota para editar usuario
# ===========================================
@app.route('/editar_usuario/<int:id>', methods=['GET', 'POST'])
def editar_usuario(id):
    conn = conectar()
    cursor = conn.cursor()

    try:
        if request.method == 'POST':
            usuario = request.form.get('usuario').strip()
            senha = request.form.get('senha', '').strip()
            nivel = request.form.get('nivel').strip().lower()

            if senha:
                senha = criptografar_senha(senha)
                cursor.execute("UPDATE usuarios SET usuario = %s, senha = %s, nivel = %s WHERE id = %s", (usuario, senha, nivel, id))
            else:
                cursor.execute("UPDATE usuarios SET usuario = %s, nivel = %s WHERE id = %s", (usuario, nivel, id))
                cursor.execute("UPDATE usuarios SET usuario = %s, nivel = %s WHERE id = %s", (usuario, nivel, id))

            conn.commit()
            return redirect(url_for('senhas'))

        cursor.execute("SELECT * FROM usuarios WHERE id = %s", (id,))
        cursor.execute("SELECT * FROM usuarios WHERE id = %s", (id,))
        usuario = cursor.fetchone()
        if not usuario:
            flash('Usuário não encontrado!', 'error')
            return redirect(url_for('senhas'))
        return render_template('editar_usuario.html', usuario=usuario)

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar usuário: {e}', 'error')
        return redirect(url_for('senhas'))

    finally:
        conn.close()



# ===========================================
# 📝 Formulário de cadastro e edição
# ===========================================
@app.route('/form')
def redirecionar_para_form():
    return redirect('/')



# ===========================================
# 📝 Rota Pagina Configurações - Ativar/Desativar
# ===========================================
@app.route('/configuracoes', methods=['GET', 'POST'])
def configuracoes_page():
    if 'usuario' not in session:
        return redirect('/login')

    if session.get('nivel') != 'admin':
        return redirect(url_for('menu', acesso_negado='1'))

    config = carregar_config()

    if request.method == 'POST':
        config['formulario_ativo'] = 'formulario_ativo' in request.form
        config['audicao_ativa'] = 'audicao_ativa' in request.form
        salvar_config(config)

    return render_template(
        'configuracoes.html',
        formulario_ativo=config['formulario_ativo'],
        audicao_ativa=config['audicao_ativa']
    )



# ===========================================
# 📝 Formulário de cadastro AUDIÇÕES
# ===========================================
@app.route('/audicoes', methods=['GET', 'POST'])
def audicoes():
    config = carregar_config()
    if not config.get('audicao_ativa', True):
        return render_template('audicoes.html', mostrar_modal=True)

    if request.method == 'POST':
        nome = request.form['nome'].strip().upper()
        telefone = request.form['telefone'].strip()
        musica = request.form['musica'].strip().title()
        link_youtube = request.form['link_youtube'].strip()
        tom = request.form['tom']
        if tom == 'Outro':
            tom = request.form.get('outro_tom').strip().upper()

        tipo = request.form['instrumento_voz']
        detalhe = ''

        if tipo == 'INSTRUMENTO' or tipo == 'AMBOS':
            instrumento = request.form.get('instrumento_escolhido')
            if instrumento == 'Outro':
                detalhe = request.form.get('outro_instrumento').strip().title()
            else:
                detalhe = instrumento
        else:
            detalhe = None

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO audicoes (nome, telefone, musica, link_youtube, tom, instrumento_voz, detalhe_instrumento, data_envio)
            VALUES (%s, %s, %s, %s, %s, %s, %s, now())
        ''', (nome, telefone, musica, link_youtube, tom, tipo, detalhe))

        conn.commit()
        cursor.close()
        conn.close()

        return render_template('sucesso.html')

    return render_template('audicoes.html')




# ===========================================
# 📝 Gerenciar Audiçoes
# ===========================================
@app.route('/audicoes_gerencia')
def audicoes_gerencia():
    if 'usuario' not in session:
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM audicoes
        WHERE arquivado = FALSE
        ORDER BY data_envio DESC
    ''')
    dados = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('audicoes_gerencia.html', dados=dados)


# ===========================================
# 📝 Thumbnail youtube
# ===========================================
@app.template_filter('youtube_id')
def youtube_id(link):
    regex = r'(?:v=|\/)([0-9A-Za-z_-]{11}).*'
    match = re.search(regex, link)
    return match.group(1) if match else ''

# ===========================================
# 📝 Filtro horario da inscrição audiçao
# ===========================================
@app.template_filter('strftime_brasil')
def _jinja2_filter_datetime_brasil(value, fmt=None):
    if not value:
        return ""
    try:
        if isinstance(value, str):
            try:
                value = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return value
        value = value - timedelta(hours=3)
        return value.strftime(fmt or "%d/%m/%y - %H:%M:%S")
    except Exception:
        return str(value)
    

# ===========================================
# 📝 Rota para excluir audição
# ===========================================
@app.route('/excluir_audicao/<int:id>')
def excluir_audicao(id):
    if 'usuario' not in session:
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM audicoes WHERE id = %s', (id,))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('audicoes_gerencia'))


# ===========================================
# 📝 Rota para arquivar audição
# ===========================================
@app.route('/arquivar_audicao/<int:id>')
def arquivar_audicao(id):
    if 'usuario' not in session:
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE audicoes SET arquivado = TRUE WHERE id = %s
    ''', (id,))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('audicoes_gerencia'))


# ===========================================
# 📝 Rota pagina audicoes arquivadas
# ===========================================
@app.route('/audicoes_arquivadas')
def audicoes_arquivadas():
    if 'usuario' not in session:
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM audicoes
        WHERE arquivado = TRUE
        ORDER BY data_envio DESC
    ''')
    dados = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('audicoes_arquivadas.html', dados=dados)


# ===========================================
# 📝 Rota para desarquivar um a audição
# ===========================================
@app.route('/desarquivar_audicao/<int:id>')
def desarquivar_audicao(id):
    if 'usuario' not in session:
        return redirect('/login')

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE audicoes SET arquivado = FALSE WHERE id = %s
    ''', (id,))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('audicoes_arquivadas'))


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
    #app.run(debug=True)
